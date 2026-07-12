from __future__ import annotations

import json
import zipfile
from importlib.util import find_spec
from pathlib import Path
from xml.etree import ElementTree


WORD_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
MAX_DOCX_XML_BYTES = 10 * 1024 * 1024
DTD_MARKERS = (b"<!DOCTYPE", b"<!ENTITY")


def read_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md", ".csv"}:
        return path.read_text(encoding="utf-8", errors="ignore")
    if suffix == ".docx":
        return read_docx_text(path)
    if suffix == ".xlsx":
        return read_xlsx_text(path)
    if suffix == ".xls":
        return "[暂不支持解析 .xls 老格式文件，请另存为 .xlsx 或 .csv 后上传。]"
    if suffix == ".json":
        return json.dumps(json.loads(path.read_text(encoding="utf-8")), ensure_ascii=False, indent=2)
    return f"[暂不支持直接解析 {suffix or '未知'} 文件，已记录文件名：{path.name}]"


def read_docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        document_info = archive.getinfo("word/document.xml")
        if document_info.file_size > MAX_DOCX_XML_BYTES:
            raise ValueError("DOCX 文档内容过大，已拒绝解析")
        document_xml = archive.read("word/document.xml")

    upper_xml = document_xml.upper()
    if any(marker in upper_xml for marker in DTD_MARKERS):
        raise ValueError("DOCX 包含不安全的 XML 实体定义，已拒绝解析")

    root = ElementTree.fromstring(document_xml)
    lines: list[str] = []

    for paragraph in root.findall(".//w:p", WORD_NS):
        parts: list[str] = []
        for text in paragraph.findall(".//w:t", WORD_NS):
            if text.text:
                parts.append(text.text)
        line = "".join(parts).strip()
        if line:
            lines.append(line)

    return "\n".join(lines)


def read_xlsx_text(path: Path) -> str:
    if not find_spec("openpyxl"):
        return "[当前环境缺少 openpyxl，无法解析 .xlsx 文件。]"

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    lines: list[str] = []

    for sheet in workbook.worksheets:
        lines.append(f"[Sheet] {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if values:
                lines.append(" | ".join(values))
            if len(lines) >= 500:
                break
        if len(lines) >= 500:
            break

    return "\n".join(lines)


def read_xlsx_rows(path: Path) -> list[dict]:
    if not find_spec("openpyxl") or path.suffix.lower() != ".xlsx":
        return []

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict] = []

    for sheet in workbook.worksheets:
        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [str(cell).strip() if cell is not None else "" for cell in row]
            if any(values):
                rows.append(
                    {
                        "sheet": sheet.title,
                        "rowNumber": index,
                        "cells": values,
                    }
                )

    return rows


def compact_lines(text: str, limit: int = 80) -> list[str]:
    lines = [" ".join(line.split()) for line in text.splitlines()]
    return [line for line in lines if line][:limit]


def keyword_hits(text: str, keywords: list[str], limit: int = 12) -> list[str]:
    hits: list[str] = []
    for line in compact_lines(text, limit=300):
        if any(keyword in line for keyword in keywords):
            hits.append(line)
        if len(hits) >= limit:
            break
    return hits
